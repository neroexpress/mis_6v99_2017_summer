import requests
import pprint
import os
import zipfile
import openpyxl
import sqlite3
import glob
import getpass
import fnmatch
import re
import csv

url = ('https://data.medicare.gov/views/bg9k-emty/files/'
      '0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2'
      'Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip')

k_url = "http://kevincrook.com/utd/hospital_ranking_focus_status.xlsx"
staging_dir_name = 'staging'
db_name = "medicare_hospital_compare.db"

def get_Medicare_Hospital_Compare_Data(staging_dir_name,url):
	r = requests.get(url)
	if (os.path.isdir(staging_dir_name)) is False:os.mkdir(staging_dir_name)
	zip_file_name = os.path.join(staging_dir_name,"test.zip")
	zf = open(zip_file_name,'wb')
	zf.write(r.content)
	zf.close()
	z = zipfile.ZipFile(zip_file_name,'r')
	z.extractall(staging_dir_name)
	z.close()
	os.remove(os.path.join(staging_dir_name,"FY2015_Percent_Change_in_Medicare_Payments.csv"))
	#print(len(fnmatch.filter(os.listdir(staging_dir_name),'*.csv')))

def get_House_Proprietary_Hospital_Rankings(k_url):
	r = requests.get(k_url)
	xf = open("hospital_ranking_focus_states.xlsx","wb")
	xf.write(r.content)
	xf.close()

def transform_name(file_name,tb):
	#print("Earlier: ",file_name)
	file_name = file_name.lower()
	file_name = re.sub(r'[\s\-\/]',"_",file_name)
	file_name = re.sub(r'%',"pct",file_name)
	if tb =='table':
		file_name = re.sub(r'^[^a-zA-Z]+',"t_",file_name)
	else:
		file_name = re.sub(r'^[^a-zA-Z]+',"c_",file_name)
	#print("Later: ",file_name)
	return file_name

def read_header(file_name):
	with open(file_name, "rt",encoding='cp1252') as f:
		d_reader = csv.DictReader(f)
		header = d_reader.fieldnames
	return header

def create_sql_table(table_name,Column_list,db_name):
	columns_tuple=tuple(Column_list)
	sql_drop_str = 'drop table if exists ' + table_name
	sql_create_str = 'create table if not exists ' + table_name  + str(columns_tuple)
	conn = sqlite3.connect(db_name)
	c1  =  conn.cursor()
	c1.execute(sql_drop_str)
	c1.execute(sql_create_str)
	c1.close()

def insert_values(file_name,table_name,Column_list,db_name):
	columns_tuple=tuple(Column_list)
	conn = sqlite3.connect(db_name)
	c1  =  conn.cursor()
	with open(file_name, "rt",encoding='cp1252') as f:
		d_reader = csv.DictReader(f)
		for line in d_reader:
			sql_tuple = tuple([line[col] for col in d_reader.fieldnames])
			blank = (sql_tuple.count(None) == len(sql_tuple)-1)
			sql_str = 'insert into ' +  table_name + str(columns_tuple) + ' values'+ str(sql_tuple)
			try:
				if blank is not True:c1.execute(sql_str)
			except Exception as e:
				raise e
	conn.commit()
	c1.close()

def creat_sqlite_db(staging_dir_name,db_name):
	glob_dir = os.path.join(staging_dir_name,"*.csv")
	for file_name in glob.glob(glob_dir):
		print(file_name)
		#print("  basename:",os.path.basename(file_name))
		header = read_header(file_name)
		#print("before: ",header)
		table_name = transform_name(os.path.splitext(os.path.basename(file_name))[0],'table')
		#print("table_name: ",table_name)
		Column_list = list()
		for head in header:
			head = transform_name(head,'column')
			Column_list.append(head)
		#print("Column_list: ",Column_list)
		create_sql_table(table_name,Column_list,db_name)
		print("Table Created: ", table_name)
		insert_values(file_name,table_name,Column_list,db_name)
		print("Values inserted: ", table_name)
		print("")
		#print("  split extension: ",os.path.splitext(os.path.basename(file_name)))
	    #print("  directory name: ", os.path.dirname(file_name))
	    #print("  absolute path: ", os.path.abspath(file_name))

def check_if_number_of_rows_matches(staging_dir_name,db_name):
	glob_dir = os.path.join(staging_dir_name,"*.csv")
	for file_name in glob.glob(glob_dir):
		with open(file_name, "rt",encoding='cp1252') as f:
			reader = csv.reader(f,delimiter = ",")
			data = list(reader)
			row_count = len(data)
			#print(row_count)
		table_name = transform_name(os.path.splitext(os.path.basename(file_name))[0],'table')
		conn = sqlite3.connect(db_name)
		c1  =  conn.cursor()
		sql_str = "select count(*) from " + table_name
		rows = c1.execute(sql_str)
		for row in rows:
		    #print(row[0])
		    pass
		numberOfRowsIntable = row[0]
		numberOfRowsInCSV = row_count-1
		print("Table: {0}, CSV: {1}".format(numberOfRowsIntable,numberOfRowsInCSV))
		if numberOfRowsIntable != numberOfRowsInCSV:
			print("Rows not equal for CSV :{0}, Table: {1}".format(file_name,table_name))
		else:print("Number of rows matches in CSV and Table")
		print("")
		c1.close()

#get_Medicare_Hospital_Compare_Data(staging_dir_name,url)
#get_House_Proprietary_Hospital_Rankings(k_url)
#creat_sqlite_db(staging_dir_name,db_name)
#check_if_number_of_rows_matches(staging_dir_name,db_name)











'''
wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")

for sheet_name in wb.get_sheet_names():
    print(sheet_name)

sheet = wb.get_sheet_by_name("Hospital National Ranking")

i= 1
while sheet.cell(row=i,column=1).value != None:
    print(sheet.cell(row=i,column=1).value, "|", sheet.cell(row=i,column=2).value)
    i +=1

sheet = wb.get_sheet_by_name("Focus States")

i= 1
while sheet.cell(row=i,column=1).value != None:
    print(sheet.cell(row=i,column=1).value, "|", sheet.cell(row=i,column=2).value)
    i +=1

wb2 = openpyxl.Workbook()

sheet_1 = wb2.create_sheet("utd")

sheet_1.cell(row=1,column=1,value="buan")

for i in range(2,11):
    sheet_1.cell(row=i,column=1,value=i-1)

sheet_2 = wb2.create_sheet("test")
sheet_2.cell(row=1,column=2,value="valued")

wb2.remove_sheet(wb2.get_sheet_by_name("Sheet"))

wb2.save("test.xlsx")

openpyxl.__version__

'''

'''

sql_str = "insert into my_table(column_1,column_2,column_3) values(?,?,?)"
sql_tuple = ('a','b','c')
c1.execute(sql_str,sql_tuple)

conn.commit()

sql_str= "select * from my_table"
rows = c1.execute(sql_str)
for row in rows:
    print(row)

sql_str = "select * from sqlite_master"
rows = c1.execute(sql_str)
for row in rows:
    print(row)

sql_str = "PRAGMA table_info('my_table')"
rows = c1.execute(sql_str)
for row in rows:
    print(row)

sql_str = "select * from sqlite_master where tbl_name = ?"
sql_tuple = ('my_table',)
c1.execute(sql_str,sql_tuple)
for row in rows:
    print(row)

sql_str = "select count(*) from my_table"
rows = c1.execute(sql_str)
for row in rows:
    print(row)

#zip_file_name = os.path.join(staging_dir_name,"test.zip")

#fn = "Timely and Effective Care - Hospital.csv"
fn= os.path.join(staging_dir_name,"Timely and Effective Care - Hospital.csv")
in_fp = open(fn,'rt',encoding='cp1252')
input_data = in_fp.read()
in_fp.close()

ofn= os.path.join(staging_dir_name,"Timely and Effective Care - Hospital.csv")
out_fp = open(ofn,'wt',encoding='utf-8')
for c in input_data:
    if c != '\0':
        out_fp.write(c)
out_fp.close()

github_username = input("Enter your github username: ")

github_password = getpass.getpass("enter your password")

github_url = "https://api.github.com/user/repos"

r = requests.get(github_url,auth=(github_username,github_password))

r.status_code

r.json()

r.headers

'''









